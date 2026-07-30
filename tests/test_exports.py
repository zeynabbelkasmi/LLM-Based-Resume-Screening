"""Tests des exports entreprise Excel et PDF sur une base temporaire."""

from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO

import pdfplumber
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.export_router import router as export_router
from backend.export_service import (
    EXCEL_MIME,
    PDF_MIME,
    EXPORT_NOTICE,
    ExportService,
    neutralize_excel_formula,
)
from backend.repository import Repository


RAW_SENTINEL = "RAW_CV_SECRET_SENTINEL_DO_NOT_EXPORT"
TOTAL_CANDIDATES = 125


def _analysis_payload(index: int) -> dict:
    name = "Élodie Dùpont – Ingénieure" if index == 2 else f"Candidat {index:03d}"
    profile = {
        "technical_skills": ["Python", "SQL", "Docker"],
        "soft_skills": ["Communication", "Rigueur"],
        "education_keywords": ["Informatique"],
        "years_experience": float(index % 12),
        "email": f"candidat{index}@example.test",
        "phone": f"+33 6 12 34 {index % 100:02d} {index % 99:02d}",
        "location": "Lyon, France",
        "headline": "Data Engineer",
    }
    sections = [
        {
            "section_name": "Compétences Techniques",
            "score": 80,
            "justification": "Python et SQL explicitement mentionnés.",
            "points_forts": ["Python"],
            "points_faibles": [],
        },
        {
            "section_name": "Soft Skills",
            "score": 70,
            "justification": "Communication mentionnée.",
            "points_forts": ["Communication"],
            "points_faibles": [],
        },
        {
            "section_name": "Formation",
            "score": 75,
            "justification": "Formation informatique.",
            "points_forts": ["Master"],
            "points_faibles": [],
        },
        {
            "section_name": "Expérience Professionnelle",
            "score": 65,
            "justification": "Expérience à confirmer en entretien.",
            "points_forts": ["Expérience mesurable"],
            "points_faibles": [],
        },
    ]
    return {
        "candidate_name": name,
        "candidate_profile": profile,
        "sections": sections,
        "global": {
            "score_global": 72.5,
            "verdict": "À CONSIDÉRER",
            "skills_presents": ["Python", "SQL"],
            "skills_absents": ["Kubernetes"],
            "forces_principales": ["Maîtrise de Python", "Bonne communication"],
            "points_amelioration": ["Kubernetes à valider"],
            "commentaire_global": "Profil cohérent ; une validation humaine reste nécessaire.",
        },
        "comparison_metrics": {"years_experience": float(index % 12)},
        "analysis_quality": "suffisante",
        "confidence": "moyenne",
    }


@pytest.fixture
def export_repository(tmp_path) -> Repository:
    repository = Repository(tmp_path / "exports.db")
    repository.init_schema()
    rows = []
    for index in range(1, TOTAL_CANDIDATES + 1):
        payload = _analysis_payload(index)
        candidate_name = payload["candidate_name"]
        status = "shortlist" if index % 3 == 0 else "nouveau"
        tags = ["Data", "Python"]
        created_at = f"2026-07-{(index % 27) + 1:02d}T12:{index % 60:02d}:00Z"
        if index == 1:
            candidate_name = '=HYPERLINK("https://example.invalid","clic")'
            status = "@commande"
            tags = ["+cmd|'/C calc'!A0", "Data"]
            created_at = "\t=NOW()"
        rows.append(
            (
                f"candidat-{index}.pdf",
                candidate_name,
                f"{RAW_SENTINEL} {index}\nEmail: candidat{index}@example.test",
                72.5,
                "À CONSIDÉRER",
                status,
                json.dumps(tags, ensure_ascii=False),
                int(index % 5 == 0),
                "Synthèse sans contenu brut.",
                json.dumps(payload, ensure_ascii=False),
                created_at,
                created_at,
            )
        )
    with repository.connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        connection.executemany(
            """
            INSERT INTO analyses (
                cv_filename, candidate_name, markdown_content, score_global,
                verdict, status, tags_json, favorite, commentaire_global,
                analysis_json, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        connection.execute("COMMIT")
    return repository


@pytest.fixture
def export_client(export_repository: Repository):
    application = FastAPI()
    application.state.repository = export_repository
    application.include_router(export_router)
    with TestClient(application) as client:
        yield client


def _pdf_text(payload: bytes) -> str:
    with pdfplumber.open(BytesIO(payload)) as document:
        return "\n".join(page.extract_text() or "" for page in document.pages)


@pytest.mark.parametrize(
    "value",
    [
        "=SUM(A1:A2)",
        "+cmd|'/C calc'!A0",
        "-2+3",
        "@SUM(A1:A2)",
        "\t=HYPERLINK(\"https://example.invalid\")",
        "\r=NOW()",
    ],
)
def test_every_excel_formula_prefix_is_neutralized(value):
    sanitized = neutralize_excel_formula(value)
    assert sanitized.startswith("'")
    assert sanitized[1:] == value


def test_excel_export_is_exhaustive_professional_and_formula_safe(export_client):
    response = export_client.get("/api/exports/candidates.xlsx")
    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == EXCEL_MIME
    assert response.content.startswith(b"PK")
    assert "attachment" in response.headers["content-disposition"]
    assert "analyse-cv-candidats-" in response.headers["content-disposition"]
    assert response.headers["cache-control"].startswith("private, no-store")

    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == ["Synthèse", "Candidats", "Détails"]
    summary = workbook["Synthèse"]
    candidates = workbook["Candidats"]
    details = workbook["Détails"]
    assert summary["B4"].value == TOTAL_CANDIDATES
    assert summary["D4"].value == "Score moyen par critère"
    assert summary["E5"].value == 80  # moyenne des scores techniques injectés
    assert candidates.max_row == TOTAL_CANDIDATES + 1
    assert candidates.freeze_panes == "A2"
    # Un Tableau Excel et un AutoFilter de feuille sur la même plage rendent le
    # fichier « à réparer » dans Excel : seul le tableau doit porter le filtre.
    assert candidates.auto_filter.ref is None
    assert "AnalyseCVCandidates" in candidates.tables
    assert candidates.tables["AnalyseCVCandidates"].ref == f"A1:AB{TOTAL_CANDIDATES + 1}"
    assert candidates["A1"].value == "ID"
    assert candidates["D1"].value == "Email"
    assert candidates["R1"].value == "Score compétences techniques"
    assert candidates["Z1"].value == "Commentaire global"
    assert candidates["AB1"].value == "Points d'amélioration"

    id_to_row = {
        candidates.cell(row, 1).value: row
        for row in range(2, candidates.max_row + 1)
    }
    injected_row = id_to_row[1]
    assert candidates.cell(injected_row, 2).value.startswith("'=")
    assert candidates.cell(injected_row, 8).value.startswith("'@")
    assert candidates.cell(injected_row, 10).value.startswith("'+")
    assert candidates.cell(injected_row, 24).value.startswith("'\t")
    assert candidates.cell(injected_row, 4).value == "candidat1@example.test"
    assert candidates.cell(injected_row, 18).value == 80

    clean_row = id_to_row[2]
    assert isinstance(candidates.cell(clean_row, 24).value, datetime)
    assert candidates.cell(clean_row, 11).number_format == "0.0"
    assert candidates.cell(clean_row, 26).value.startswith("Profil cohérent")

    # Détails : une ligne par candidat et par critère, avec justification.
    assert details.max_row == TOTAL_CANDIDATES * 4 + 1
    assert details["C1"].value == "Critère"
    assert "AnalyseCVDetails" in details.tables
    detail_justifications = {
        details.cell(row, 5).value for row in range(2, details.max_row + 1)
    }
    assert "Python et SQL explicitement mentionnés." in detail_justifications
    workbook.close()


def test_global_pdf_contains_all_candidates_and_never_raw_cv(export_client):
    response = export_client.get("/api/exports/candidates.pdf")
    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == PDF_MIME
    assert response.content.startswith(b"%PDF-")
    assert "analyse-cv-candidats-" in response.headers["content-disposition"]

    text = _pdf_text(response.content)
    assert "Rapport consolidé des candidats" in text
    assert f"CANDIDATS\n{TOTAL_CANDIDATES}" in text or str(TOTAL_CANDIDATES) in text
    assert "Candidat 125" in text
    assert "validation" in text.lower()
    assert "confidentiel" in text.lower()
    assert RAW_SENTINEL not in text
    assert RAW_SENTINEL.encode() not in response.content


def test_individual_pdf_is_unicode_detailed_and_excludes_raw_document(export_client):
    response = export_client.get("/api/exports/candidates/2.pdf")
    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == PDF_MIME
    assert response.content.startswith(b"%PDF-")
    assert 'filename="analyse-cv-candidat-2.pdf"' in response.headers["content-disposition"]

    text = _pdf_text(response.content)
    assert "Élodie Dùpont" in text
    assert "candidat2@example.test" in text
    assert "Compétences Techniques" in text
    assert "Python et SQL explicitement mentionnés" in text
    assert "texte brut du CV" in text
    assert RAW_SENTINEL not in text


def test_individual_pdf_returns_404_for_unknown_candidate(export_client):
    response = export_client.get("/api/exports/candidates/99999.pdf")
    assert response.status_code == 404
    assert response.json() == {"detail": "Candidat introuvable"}


def test_export_audit_hook_contains_no_candidate_personal_data(export_repository):
    events: list[dict] = []
    service = ExportService(export_repository, audit_hook=events.append)
    payload = service.candidate_pdf(2)
    assert payload and payload.startswith(b"%PDF-")
    assert events == [
        {
            "event": "candidate_export_generated",
            "format": "pdf_detail",
            "candidate_count": 1,
            "analysis_id": 2,
            "generated_at": events[0]["generated_at"],
        }
    ]
    assert "Élodie" not in json.dumps(events, ensure_ascii=False)
    assert EXPORT_NOTICE not in json.dumps(events, ensure_ascii=False)
