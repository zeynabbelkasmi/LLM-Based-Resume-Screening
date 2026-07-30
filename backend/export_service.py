"""Exports RH professionnels, exhaustifs et sans contenu de CV brut par défaut.

Le service reste indépendant de FastAPI afin de pouvoir être utilisé depuis une
tâche planifiée, une CLI ou les routes HTTP. Les documents sont parcourus avec
``Repository.iter_analysis_documents`` : aucun plafond de pagination de l'API ne
s'applique aux exports.
"""

from __future__ import annotations

import html
import logging
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from statistics import fmean
from typing import Any, Callable, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from reportlab import __file__ as reportlab_file
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .repository import Repository


logger = logging.getLogger(__name__)

EXPORT_NOTICE = (
    "Document confidentiel. Les scores sont des aides à l'analyse et ne doivent "
    "jamais remplacer la validation d'un recruteur habilité."
)
EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"

AuditHook = Callable[[dict[str, Any]], None]


@dataclass(slots=True)
class CandidateExport:
    """Vue d'export volontairement séparée du CV brut et des notes internes."""

    analysis_id: int
    name: str
    source_filename: str
    email: str = ""
    phone: str = ""
    location: str = ""
    headline: str = ""
    status: str = "nouveau"
    favorite: bool = False
    tags: list[str] = field(default_factory=list)
    score: float = 0.0
    verdict: str = ""
    years_experience: float = 0.0
    matched_skills: list[str] = field(default_factory=list)
    technical_skills: list[str] = field(default_factory=list)
    soft_skills: list[str] = field(default_factory=list)
    missing_skills: list[str] = field(default_factory=list)
    section_scores: dict[str, float] = field(default_factory=dict)
    section_details: list[dict[str, Any]] = field(default_factory=list)
    created_at: str = ""
    updated_at: str = ""
    analysis_quality: str = ""
    confidence: str = ""
    global_comment: str = ""
    strengths: list[str] = field(default_factory=list)
    improvements: list[str] = field(default_factory=list)


def _generated_at() -> datetime:
    return datetime.now(UTC).replace(microsecond=0)


def _as_mapping(value: Any) -> Mapping[str, Any]:
    return value if isinstance(value, Mapping) else {}


def _as_strings(value: Any, *, limit: int = 100) -> list[str]:
    if not isinstance(value, (list, tuple, set)):
        return []
    result: list[str] = []
    seen: set[str] = set()
    for item in value:
        text = str(item or "").strip()
        key = text.casefold()
        if text and key not in seen:
            result.append(text[:500])
            seen.add(key)
        if len(result) >= limit:
            break
    return result


def _number(value: Any) -> float:
    try:
        number = float(value or 0)
    except (TypeError, ValueError, OverflowError):
        return 0.0
    if number != number or number in (float("inf"), float("-inf")):
        return 0.0
    return number


def _fallback_name(filename: str) -> str:
    stem = Path(filename or "candidat").stem
    name = re.sub(r"[_-]+", " ", stem).strip()
    return name.title() or "Candidat"


def _extract_contacts(markdown: str, profile: Mapping[str, Any]) -> dict[str, str]:
    """Complète les anciennes analyses sans conserver le texte source."""

    text = str(markdown or "")
    email = str(profile.get("email") or "").strip()
    phone = str(profile.get("phone") or "").strip()
    location = str(profile.get("location") or "").strip()
    headline = str(profile.get("headline") or "").strip()
    if not email:
        match = re.search(
            r"(?<![\w.+-])[\w.+-]+@[\w-]+(?:\.[\w-]+)+(?![\w.-])",
            text,
            flags=re.IGNORECASE,
        )
        email = match.group(0) if match else ""
    if not phone:
        match = re.search(
            r"(?<!\w)(?:\+\d{1,3}[\s().-]*)?(?:\d[\s().-]*){8,14}\d(?!\w)",
            text,
        )
        phone = match.group(0).strip(" .-") if match else ""
    if not location or not headline:
        for raw_line in text.splitlines()[:120]:
            line = raw_line.strip().lstrip("#-* ").replace("**", "").strip()
            if not location:
                match = re.match(
                    r"(?:localisation|ville|adresse|location)\s*[:\-–—]\s*(.+)$",
                    line,
                    flags=re.IGNORECASE,
                )
                location = match.group(1).strip() if match else ""
            if not headline:
                match = re.match(
                    r"(?:poste(?:\s+recherché)?|titre|profil)\s*[:\-–—]\s*(.+)$",
                    line,
                    flags=re.IGNORECASE,
                )
                headline = match.group(1).strip() if match else ""
            if location and headline:
                break
    return {
        "email": email[:255],
        "phone": phone[:80],
        "location": location[:255],
        "headline": headline[:255],
    }


def _candidate_from_document(document: Mapping[str, Any]) -> CandidateExport:
    analysis = _as_mapping(document.get("analysis"))
    global_data = _as_mapping(analysis.get("global"))
    profile = _as_mapping(analysis.get("candidate_profile"))
    metrics = _as_mapping(analysis.get("comparison_metrics"))
    contacts = _extract_contacts(str(document.get("markdown_content") or ""), profile)

    raw_sections = analysis.get("sections")
    sections = [dict(item) for item in raw_sections if isinstance(item, Mapping)] if isinstance(raw_sections, list) else []
    section_scores: dict[str, float] = {}
    for section in sections:
        name = str(section.get("section_name") or section.get("nom") or "").strip()
        if name:
            section_scores[name] = _number(section.get("score", section.get("score_sur_100")))

    filename = str(document.get("cv_filename") or "candidat")
    return CandidateExport(
        analysis_id=int(document.get("id") or 0),
        name=str(document.get("candidate_name") or "").strip() or _fallback_name(filename),
        source_filename=filename,
        email=contacts["email"],
        phone=contacts["phone"],
        location=contacts["location"],
        headline=contacts["headline"],
        status=str(document.get("status") or "nouveau"),
        favorite=bool(document.get("favorite")),
        tags=_as_strings(document.get("tags")),
        score=_number(document.get("score_global", global_data.get("score_global"))),
        verdict=str(document.get("verdict") or global_data.get("verdict") or ""),
        years_experience=_number(metrics.get("years_experience", profile.get("years_experience"))),
        matched_skills=_as_strings(global_data.get("skills_presents")),
        technical_skills=_as_strings(profile.get("technical_skills")),
        soft_skills=_as_strings(profile.get("soft_skills")),
        missing_skills=_as_strings(global_data.get("skills_absents")),
        section_scores=section_scores,
        section_details=sections,
        created_at=str(document.get("created_at") or ""),
        updated_at=str(document.get("updated_at") or document.get("created_at") or ""),
        analysis_quality=str(analysis.get("analysis_quality") or global_data.get("analysis_quality") or ""),
        confidence=str(analysis.get("confidence") or global_data.get("confidence") or ""),
        global_comment=str(global_data.get("commentaire_global") or document.get("commentaire_global") or ""),
        strengths=_as_strings(global_data.get("forces_principales")),
        improvements=_as_strings(global_data.get("points_amelioration")),
    )


def _section_score(candidate: CandidateExport, label: str) -> float:
    wanted = re.sub(r"[^a-z]", "", label.casefold())
    for name, value in candidate.section_scores.items():
        normalized = re.sub(r"[^a-z]", "", name.casefold())
        if normalized == wanted or wanted in normalized or normalized in wanted:
            return value
    return 0.0


def neutralize_excel_formula(value: Any) -> Any:
    """Empêche l'exécution d'une formule injectée dans une cellule texte."""

    if not isinstance(value, str) or not value:
        return value
    # Excel refuse certains contrôles XML et limite une cellule à 32 767
    # caractères. Les tabulations/retours sont conservés mais neutralisés en
    # première position, car Excel les utilise pour contourner les protections
    # naïves contre l'injection de formules.
    cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)[:32767]
    if not cleaned:
        return ""
    first_visible = cleaned.lstrip(" \ufeff")[:1]
    if cleaned[0] in "\t\r\n" or first_visible in {"=", "+", "-", "@", "\t", "\r", "\n"}:
        return "'" + cleaned[:32766]
    return cleaned


def _join(values: Sequence[str]) -> str:
    return " • ".join(values)


def _as_excel_datetime(value: Any) -> datetime | None:
    """Excel exige des datetimes naïfs ; toute valeur douteuse reste du texte."""

    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(UTC).replace(tzinfo=None)
    return parsed


def _status_label(value: str) -> str:
    labels = {
        "nouveau": "Nouveau",
        "a_revoir": "À revoir",
        "a_contacter": "À contacter",
        "entretien": "Entretien",
        "retenu": "Retenu",
        "shortlist": "Shortlist",
        "refuse": "Refusé",
        "embauche": "Embauché",
    }
    return labels.get(value.casefold(), value.replace("_", " ").strip().title())


class ExportService:
    """Produit les exports et publie un événement d'audit sans données candidat."""

    def __init__(self, repository: Repository, audit_hook: AuditHook | None = None):
        self.repository = repository
        self.audit_hook = audit_hook

    def all_candidates(self) -> list[CandidateExport]:
        candidates: list[CandidateExport] = []
        for document in self.repository.iter_analysis_documents(batch_size=100):
            candidate = _candidate_from_document(document)
            # Les anciennes bases peuvent avoir les sections uniquement dans la table
            # relationnelle. On ne fait une requête supplémentaire que pour ces lignes.
            if not candidate.section_details:
                detail = self.repository.get_analysis(candidate.analysis_id, include_document=False)
                if detail:
                    detail_sections = detail.get("sections") or []
                    candidate.section_details = [
                        dict(item) for item in detail_sections if isinstance(item, Mapping)
                    ]
                    for section in candidate.section_details:
                        name = str(section.get("section_name") or "").strip()
                        if name:
                            candidate.section_scores[name] = _number(section.get("score"))
            candidates.append(candidate)
        return candidates

    def candidate(self, analysis_id: int) -> CandidateExport | None:
        document = self.repository.get_analysis(analysis_id, include_document=True)
        if document is None:
            return None
        # ``get_analysis`` nomme le JSON structuré ``analysis`` comme l'itérateur.
        return _candidate_from_document(document)

    def _publish_audit(
        self, *, export_format: str, candidate_count: int, analysis_id: int | None = None
    ) -> None:
        event = {
            "event": "candidate_export_generated",
            "format": export_format,
            "candidate_count": candidate_count,
            "analysis_id": analysis_id,
            "generated_at": _generated_at().isoformat().replace("+00:00", "Z"),
        }
        logger.info("Export RH généré", extra={"export_event": event})
        if self.audit_hook:
            try:
                self.audit_hook(event)
            except Exception:  # pragma: no cover - un audit externe ne bloque pas l'export
                logger.exception("Le hook d'audit d'export a échoué")

    def candidates_xlsx(self) -> bytes:
        candidates = self.all_candidates()
        payload = _build_candidates_workbook(candidates)
        self._publish_audit(export_format="xlsx", candidate_count=len(candidates))
        return payload

    def candidates_pdf(self) -> bytes:
        candidates = self.all_candidates()
        payload = _build_candidates_pdf(candidates)
        self._publish_audit(export_format="pdf", candidate_count=len(candidates))
        return payload

    def candidate_pdf(self, analysis_id: int) -> bytes | None:
        candidate = self.candidate(analysis_id)
        if candidate is None:
            return None
        payload = _build_candidate_pdf(candidate)
        self._publish_audit(
            export_format="pdf_detail", candidate_count=1, analysis_id=analysis_id
        )
        return payload


_CRITERION_COLUMNS = (
    ("Compétences techniques", "Compétences Techniques"),
    ("Soft skills", "Soft Skills"),
    ("Formation", "Formation"),
    ("Expérience", "Expérience Professionnelle"),
)


def _criterion_average(candidates: Sequence[CandidateExport], label: str) -> float | None:
    """Moyenne d'un critère sur les seuls candidats qui possèdent ce score."""

    wanted = re.sub(r"[^a-z]", "", label.casefold())
    values: list[float] = []
    for candidate in candidates:
        for name, score in candidate.section_scores.items():
            normalized = re.sub(r"[^a-z]", "", name.casefold())
            if normalized == wanted or wanted in normalized or normalized in wanted:
                values.append(score)
                break
    return round(fmean(values), 1) if values else None


def _build_candidates_workbook(candidates: Sequence[CandidateExport]) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Synthèse"
    worksheet = workbook.create_sheet("Candidats")
    details = workbook.create_sheet("Détails")
    workbook.properties.title = "Analyse CV — Export candidats"
    workbook.properties.subject = "Synthèse confidentielle de la CVthèque"
    workbook.properties.creator = "Analyse CV"
    workbook.calculation.fullCalcOnLoad = False

    navy = "0B1220"
    teal = "14B8A6"
    pale_teal = "CCFBF1"
    light = "F4F7FB"
    slate = "475569"
    white = "FFFFFF"
    border = Border(bottom=Side(style="thin", color="D8E1EC"))
    date_format = "DD/MM/YYYY HH:MM"

    # Un classement lisible : meilleurs scores d'abord, ordre stable par nom.
    ranked = sorted(candidates, key=lambda item: (-item.score, item.name.casefold()))

    # ── Feuille Synthèse ─────────────────────────────────────────────────────
    summary.merge_cells("A1:F2")
    title_cell = summary["A1"]
    title_cell.value = "ANALYSE CV  |  EXPORT CANDIDATS"
    title_cell.font = Font(name="Aptos Display", size=22, bold=True, color=white)
    title_cell.fill = PatternFill("solid", fgColor=navy)
    title_cell.alignment = Alignment(vertical="center")
    for row in summary["A1:F2"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=navy)

    generated = _generated_at()
    scores = [candidate.score for candidate in candidates]
    summary_metrics = [
        ("Candidats exportés", len(candidates)),
        ("Score moyen", round(fmean(scores), 2) if scores else 0),
        ("Favoris", sum(candidate.favorite for candidate in candidates)),
        ("Date de génération (UTC)", generated.strftime("%d/%m/%Y %H:%M")),
    ]
    for index, (label, value) in enumerate(summary_metrics, start=4):
        summary.cell(index, 1, label).font = Font(bold=True, color=slate)
        summary.cell(index, 2, value).font = Font(size=13, bold=True, color=navy)
        summary.cell(index, 1).fill = summary.cell(index, 2).fill = PatternFill(
            "solid", fgColor=light
        )

    average_title = summary.cell(4, 4, "Score moyen par critère")
    average_title.font = Font(size=12, bold=True, color=white)
    average_title.fill = PatternFill("solid", fgColor=teal)
    summary.cell(4, 5).fill = PatternFill("solid", fgColor=teal)
    for offset, (short_label, canonical) in enumerate(_CRITERION_COLUMNS, start=5):
        summary.cell(offset, 4, short_label).font = Font(bold=True, color=slate)
        average = _criterion_average(candidates, canonical)
        value_cell = summary.cell(offset, 5, average if average is not None else "—")
        value_cell.font = Font(size=12, bold=True, color=navy)
        value_cell.number_format = "0.0"
        summary.cell(offset, 4).fill = summary.cell(offset, 5).fill = PatternFill(
            "solid", fgColor=light
        )

    status_counts = Counter(_status_label(item.status) for item in candidates)
    verdict_counts = Counter(item.verdict or "Non défini" for item in candidates)
    summary["A10"] = "Répartition par statut"
    summary["D10"] = "Répartition par verdict"
    for coordinate in ("A10", "D10"):
        summary[coordinate].font = Font(size=12, bold=True, color=white)
        summary[coordinate].fill = PatternFill("solid", fgColor=teal)
    for row_index, (label, count) in enumerate(status_counts.most_common(), start=11):
        summary.cell(row_index, 1, neutralize_excel_formula(label))
        summary.cell(row_index, 2, count)
    for row_index, (label, count) in enumerate(verdict_counts.most_common(), start=11):
        summary.cell(row_index, 4, neutralize_excel_formula(label))
        summary.cell(row_index, 5, count)

    notice_row = max(18, 12 + max(len(status_counts), len(verdict_counts)))
    summary.merge_cells(start_row=notice_row, start_column=1, end_row=notice_row + 2, end_column=6)
    notice = summary.cell(notice_row, 1, EXPORT_NOTICE)
    notice.font = Font(italic=True, color=slate)
    notice.fill = PatternFill("solid", fgColor=pale_teal)
    notice.alignment = Alignment(wrap_text=True, vertical="center")
    summary.freeze_panes = "A4"
    summary.sheet_view.showGridLines = False
    for column, width in {"A": 31, "B": 18, "C": 4, "D": 31, "E": 18, "F": 4}.items():
        summary.column_dimensions[column].width = width

    # ── Feuille Candidats ────────────────────────────────────────────────────
    headers = [
        "ID",
        "Nom du candidat",
        "Fichier source",
        "Email",
        "Téléphone",
        "Localisation",
        "Titre professionnel",
        "Statut",
        "Favori",
        "Tags",
        "Score global / 100",
        "Verdict",
        "Années d'expérience",
        "Compétences correspondantes",
        "Compétences techniques détectées",
        "Soft skills détectées",
        "Compétences manquantes",
        "Score compétences techniques",
        "Score soft skills",
        "Score formation",
        "Score expérience professionnelle",
        "Qualité d'analyse",
        "Confiance",
        "Date d'analyse",
        "Dernière mise à jour",
        "Commentaire global",
        "Forces principales",
        "Points d'amélioration",
    ]
    last_column = get_column_letter(len(headers))
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(name="Aptos", size=10, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[1].height = 36

    for candidate in ranked:
        row = [
            candidate.analysis_id,
            candidate.name,
            candidate.source_filename,
            candidate.email,
            candidate.phone,
            candidate.location,
            candidate.headline,
            _status_label(candidate.status),
            "Oui" if candidate.favorite else "Non",
            _join(candidate.tags),
            round(candidate.score, 2),
            candidate.verdict,
            round(candidate.years_experience, 1),
            _join(candidate.matched_skills),
            _join(candidate.technical_skills),
            _join(candidate.soft_skills),
            _join(candidate.missing_skills),
            round(_section_score(candidate, "Compétences Techniques"), 2),
            round(_section_score(candidate, "Soft Skills"), 2),
            round(_section_score(candidate, "Formation"), 2),
            round(_section_score(candidate, "Expérience Professionnelle"), 2),
            candidate.analysis_quality,
            candidate.confidence,
            _as_excel_datetime(candidate.created_at) or candidate.created_at,
            _as_excel_datetime(candidate.updated_at) or candidate.updated_at,
            candidate.global_comment,
            _join(candidate.strengths),
            _join(candidate.improvements),
        ]
        worksheet.append([neutralize_excel_formula(value) for value in row])

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    # Formats homogènes : scores à une décimale, dates réelles filtrables.
    number_formats = {"K": "0.0", "M": "0.0", "R": "0.0", "S": "0.0", "T": "0.0", "U": "0.0",
                      "X": date_format, "Y": date_format}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            wanted_format = number_formats.get(cell.column_letter)
            if wanted_format:
                cell.number_format = wanted_format
        # Hauteur uniforme : le contenu long reste accessible sans lignes géantes.
        worksheet.row_dimensions[row[0].row].height = 42

    if candidates:
        # Un Tableau Excel fournit déjà tri et filtres ; ajouter en plus un
        # AutoFilter de feuille sur la même plage rend le fichier « à réparer ».
        table = ExcelTable(
            displayName="AnalyseCVCandidates",
            ref=f"A1:{last_column}{worksheet.max_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        worksheet.conditional_formatting.add(
            f"K2:K{worksheet.max_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="FECACA",
                mid_type="num",
                mid_value=60,
                mid_color="FEF3C7",
                end_type="num",
                end_value=100,
                end_color="A7F3D0",
            ),
        )
        for column in ("R", "S", "T", "U"):
            worksheet.conditional_formatting.add(
                f"{column}2:{column}{worksheet.max_row}",
                DataBarRule(
                    start_type="num",
                    start_value=0,
                    end_type="num",
                    end_value=100,
                    color=teal,
                    showValue=True,
                    minLength=0,
                    maxLength=100,
                ),
            )

    widths = [10, 26, 24, 29, 20, 24, 29, 18, 10, 26, 18, 21, 19, 37, 42, 34, 38, 20, 18, 16, 24, 19, 15, 19, 19, 46, 40, 40]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    # ── Feuille Détails : justifications par critère ─────────────────────────
    detail_headers = [
        "ID",
        "Candidat",
        "Critère",
        "Score /100",
        "Justification",
        "Points forts",
        "Points faibles",
    ]
    details.append(detail_headers)
    for cell in details[1]:
        cell.font = Font(name="Aptos", size=10, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    details.row_dimensions[1].height = 30

    for candidate in ranked:
        for section in candidate.section_details:
            name = str(section.get("section_name") or section.get("nom") or "").strip()
            if not name:
                continue
            detail_row = [
                candidate.analysis_id,
                candidate.name,
                name,
                round(_number(section.get("score", section.get("score_sur_100"))), 2),
                str(section.get("justification") or "").strip(),
                _join(_as_strings(section.get("points_forts"), limit=10)),
                _join(_as_strings(section.get("points_faibles"), limit=10)),
            ]
            details.append([neutralize_excel_formula(value) for value in detail_row])

    details.freeze_panes = "A2"
    details.sheet_view.showGridLines = False
    for row in details.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        row[3].number_format = "0.0"
    if details.max_row > 1:
        detail_table = ExcelTable(
            displayName="AnalyseCVDetails",
            ref=f"A1:{get_column_letter(len(detail_headers))}{details.max_row}",
        )
        detail_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        details.add_table(detail_table)
    for column, width in zip("ABCDEFG", (10, 26, 26, 12, 62, 40, 40)):
        details.column_dimensions[column].width = width

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _register_pdf_fonts() -> tuple[str, str]:
    """Utilise les polices Vera fournies avec ReportLab, sans appel réseau."""

    regular_name = "AnalyseCVVera"
    bold_name = "AnalyseCVVeraBold"
    if regular_name in pdfmetrics.getRegisteredFontNames():
        return regular_name, bold_name
    font_dir = Path(reportlab_file).resolve().parent / "fonts"
    try:
        pdfmetrics.registerFont(TTFont(regular_name, str(font_dir / "Vera.ttf")))
        pdfmetrics.registerFont(TTFont(bold_name, str(font_dir / "VeraBd.ttf")))
        pdfmetrics.registerFontFamily(
            regular_name,
            normal=regular_name,
            bold=bold_name,
            italic=regular_name,
            boldItalic=bold_name,
        )
        return regular_name, bold_name
    except Exception:  # pragma: no cover - fallback pour une distribution atypique
        logger.warning("Polices Vera indisponibles, utilisation du fallback PDF standard")
        return "Helvetica", "Helvetica-Bold"


def _pdf_styles() -> dict[str, ParagraphStyle]:
    regular, bold = _register_pdf_fonts()
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "AnalyseCVTitle",
            parent=base["Title"],
            fontName=bold,
            fontSize=25,
            leading=31,
            textColor=colors.HexColor("#0B1220"),
            alignment=TA_LEFT,
            spaceAfter=10,
        ),
        "subtitle": ParagraphStyle(
            "AnalyseCVSubtitle",
            parent=base["Normal"],
            fontName=regular,
            fontSize=10,
            leading=15,
            textColor=colors.HexColor("#475569"),
            spaceAfter=8,
        ),
        "heading": ParagraphStyle(
            "AnalyseCVHeading",
            parent=base["Heading2"],
            fontName=bold,
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#0F766E"),
            spaceBefore=8,
            spaceAfter=7,
        ),
        "body": ParagraphStyle(
            "AnalyseCVBody",
            parent=base["BodyText"],
            fontName=regular,
            fontSize=9,
            leading=13,
            textColor=colors.HexColor("#1E293B"),
            spaceAfter=5,
        ),
        "small": ParagraphStyle(
            "AnalyseCVSmall",
            parent=base["BodyText"],
            fontName=regular,
            fontSize=7.2,
            leading=9.2,
            textColor=colors.HexColor("#334155"),
        ),
        "small_bold": ParagraphStyle(
            "AnalyseCVSmallBold",
            parent=base["BodyText"],
            fontName=bold,
            fontSize=7.2,
            leading=9.2,
            textColor=colors.white,
            alignment=TA_CENTER,
        ),
        "notice": ParagraphStyle(
            "AnalyseCVNotice",
            parent=base["BodyText"],
            fontName=regular,
            fontSize=8.5,
            leading=12,
            textColor=colors.HexColor("#115E59"),
            borderColor=colors.HexColor("#5EEAD4"),
            borderWidth=0.7,
            borderPadding=8,
            backColor=colors.HexColor("#F0FDFA"),
        ),
        "center": ParagraphStyle(
            "AnalyseCVCenter",
            parent=base["BodyText"],
            fontName=regular,
            fontSize=9,
            leading=13,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#475569"),
        ),
        "metric": ParagraphStyle(
            "AnalyseCVMetric",
            parent=base["BodyText"],
            fontName=bold,
            fontSize=17,
            leading=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#0B1220"),
        ),
        "metric_label": ParagraphStyle(
            "AnalyseCVMetricLabel",
            parent=base["BodyText"],
            fontName=regular,
            fontSize=7,
            leading=9,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#64748B"),
        ),
    }


def _p(value: Any, style: ParagraphStyle, *, empty: str = "—") -> Paragraph:
    text = str(value or "").strip() or empty
    text = "".join(character for character in text if character in "\n\t" or ord(character) >= 32)
    return Paragraph(html.escape(text).replace("\n", "<br/>"), style)


def _page_decorator(canvas, doc) -> None:
    regular, bold = _register_pdf_fonts()
    canvas.saveState()
    width, height = doc.pagesize
    canvas.setStrokeColor(colors.HexColor("#CBD5E1"))
    canvas.setLineWidth(0.4)
    canvas.line(doc.leftMargin, 13 * mm, width - doc.rightMargin, 13 * mm)
    canvas.setFont(bold, 7)
    canvas.setFillColor(colors.HexColor("#0F766E"))
    canvas.drawString(doc.leftMargin, height - 10 * mm, "ANALYSE CV")
    canvas.setFont(regular, 7)
    canvas.setFillColor(colors.HexColor("#64748B"))
    canvas.drawRightString(width - doc.rightMargin, 8 * mm, f"Confidentiel  •  Page {doc.page}")
    canvas.restoreState()


def _metric_table(candidates: Sequence[CandidateExport], styles) -> Table:
    scores = [candidate.score for candidate in candidates]
    values = [
        ("CANDIDATS", str(len(candidates))),
        ("SCORE MOYEN", f"{fmean(scores):.1f}/100" if scores else "0/100"),
        ("FAVORIS", str(sum(candidate.favorite for candidate in candidates))),
    ]
    value_cells = [_p(value, styles["metric"]) for _, value in values]
    label_cells = [_p(label, styles["metric_label"]) for label, _ in values]
    table = Table(
        [value_cells, label_cells],
        colWidths=[53 * mm] * 3,
        rowHeights=[15 * mm, 8 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F1F5F9")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.7, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return table


def _build_candidates_pdf(candidates: Sequence[CandidateExport]) -> bytes:
    stream = BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title="Analyse CV — Export candidats",
        author="Analyse CV",
        subject="Synthèse confidentielle de la CVthèque",
    )
    styles = _pdf_styles()
    generated = _generated_at()
    story: list[Any] = [
        Spacer(1, 15 * mm),
        _p("ANALYSE CV", styles["subtitle"]),
        _p("Rapport consolidé des candidats", styles["title"]),
        _p(
            f"Photographie exhaustive de la CVthèque au {generated.strftime('%d/%m/%Y à %H:%M')} UTC.",
            styles["subtitle"],
        ),
        Spacer(1, 8 * mm),
        _metric_table(candidates, styles),
        Spacer(1, 10 * mm),
    ]

    status_counts = Counter(_status_label(candidate.status) for candidate in candidates)
    verdict_counts = Counter(candidate.verdict or "Non défini" for candidate in candidates)
    distribution = Table(
        [
            [_p("STATUTS", styles["small_bold"]), _p("VERDICTS", styles["small_bold"])],
            [
                _p("  •  ".join(f"{key}: {value}" for key, value in status_counts.most_common()) or "Aucune donnée", styles["body"]),
                _p("  •  ".join(f"{key}: {value}" for key, value in verdict_counts.most_common()) or "Aucune donnée", styles["body"]),
            ],
        ],
        colWidths=[80 * mm, 80 * mm],
    )
    distribution.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.extend([distribution, Spacer(1, 10 * mm), _p(EXPORT_NOTICE, styles["notice"]), PageBreak()])
    story.extend([_p("Vue consolidée", styles["title"]), _p("Les données sont triées par date d'analyse décroissante. Le texte brut des CV et les notes internes sont exclus.", styles["subtitle"]), Spacer(1, 3 * mm)])

    header = ["ID", "Candidat", "Statut", "Score", "Verdict", "Compétences correspondantes", "Expérience", "Date"]
    rows: list[list[Any]] = [[_p(item, styles["small_bold"]) for item in header]]
    for candidate in candidates:
        rows.append(
            [
                _p(candidate.analysis_id, styles["small"]),
                _p(candidate.name, styles["small"]),
                _p(_status_label(candidate.status), styles["small"]),
                _p(f"{candidate.score:.1f}", styles["small"]),
                _p(candidate.verdict, styles["small"]),
                _p(_join(candidate.matched_skills[:8]), styles["small"]),
                _p(f"{candidate.years_experience:g} an(s)", styles["small"]),
                _p(candidate.created_at[:10], styles["small"]),
            ]
        )
    if len(rows) == 1:
        rows.append([_p("Aucun candidat", styles["small"])] + [""] * 7)
    table = Table(
        rows,
        colWidths=[12 * mm, 39 * mm, 26 * mm, 16 * mm, 31 * mm, 76 * mm, 24 * mm, 25 * mm],
        repeatRows=1,
        splitByRow=True,
        hAlign="LEFT",
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B1220")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([table, Spacer(1, 6 * mm), _p(EXPORT_NOTICE, styles["notice"])])
    document.build(story, onFirstPage=_page_decorator, onLaterPages=_page_decorator)
    return stream.getvalue()


def _bullet_block(values: Sequence[str], styles) -> list[Any]:
    if not values:
        return [_p("—", styles["body"])]
    return [_p(f"• {value}", styles["body"]) for value in values]


def _build_candidate_pdf(candidate: CandidateExport) -> bytes:
    stream = BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=A4,
        rightMargin=17 * mm,
        leftMargin=17 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Analyse CV — {candidate.name}",
        author="Analyse CV",
        subject="Rapport individuel confidentiel",
    )
    styles = _pdf_styles()
    story: list[Any] = [
        _p("ANALYSE CV  |  RAPPORT INDIVIDUEL", styles["subtitle"]),
        _p(candidate.name, styles["title"]),
    ]
    identity_data = [
        [_p("Statut", styles["small_bold"]), _p("Score", styles["small_bold"]), _p("Verdict", styles["small_bold"])],
        [_p(_status_label(candidate.status), styles["body"]), _p(f"{candidate.score:.1f}/100", styles["body"]), _p(candidate.verdict, styles["body"])],
    ]
    identity = Table(identity_data, colWidths=[50 * mm, 35 * mm, 87 * mm])
    identity.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F0FDFA")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#99F6E4")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend([identity, Spacer(1, 5 * mm), _p("Coordonnées et profil", styles["heading"])])
    contact_rows = [
        ["Email", candidate.email],
        ["Téléphone", candidate.phone],
        ["Localisation", candidate.location],
        ["Titre", candidate.headline],
        ["Expérience détectée", f"{candidate.years_experience:g} an(s)"],
        ["Tags", _join(candidate.tags)],
        ["Date d'analyse", candidate.created_at],
        ["Dernière mise à jour", candidate.updated_at],
    ]
    contacts = Table(
        [[_p(label, styles["body"]), _p(value, styles["body"])] for label, value in contact_rows],
        colWidths=[52 * mm, 120 * mm],
    )
    contacts.setStyle(
        TableStyle(
            [
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([contacts, Spacer(1, 4 * mm), _p("Compétences", styles["heading"])])
    skill_table = Table(
        [
            [_p("Correspondances", styles["small_bold"]), _p("Techniques détectées", styles["small_bold"]), _p("Manquantes", styles["small_bold"])],
            [_p(_join(candidate.matched_skills), styles["body"]), _p(_join(candidate.technical_skills), styles["body"]), _p(_join(candidate.missing_skills), styles["body"])],
        ],
        colWidths=[57 * mm, 57 * mm, 58 * mm],
    )
    skill_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B1220")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend([skill_table, Spacer(1, 4 * mm), _p("Scores détaillés", styles["heading"])])

    section_rows: list[list[Any]] = [[_p("Critère", styles["small_bold"]), _p("Score", styles["small_bold"]), _p("Justification", styles["small_bold"])]]
    for section in candidate.section_details:
        section_rows.append(
            [
                _p(section.get("section_name") or section.get("nom"), styles["small"]),
                _p(f"{_number(section.get('score', section.get('score_sur_100'))):.1f}/100", styles["small"]),
                _p(section.get("justification"), styles["small"]),
            ]
        )
    if len(section_rows) == 1:
        section_rows.append([_p("Aucun score détaillé", styles["small"]), "", ""])
    section_table = Table(section_rows, colWidths=[45 * mm, 24 * mm, 103 * mm], repeatRows=1)
    section_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([section_table, Spacer(1, 4 * mm)])

    summary_parts: list[Any] = [_p("Synthèse d'aide à la décision", styles["heading"])]
    if candidate.global_comment:
        summary_parts.append(_p(candidate.global_comment, styles["body"]))
    summary_parts.extend([_p("Points forts", styles["heading"]), *_bullet_block(candidate.strengths, styles)])
    summary_parts.extend([_p("Points à vérifier", styles["heading"]), *_bullet_block(candidate.improvements, styles)])
    story.extend([KeepTogether(summary_parts), Spacer(1, 5 * mm), _p(EXPORT_NOTICE, styles["notice"]), Spacer(1, 2 * mm), _p("Le texte brut du CV et les notes internes ne sont pas inclus dans ce rapport.", styles["subtitle"])])
    document.build(story, onFirstPage=_page_decorator, onLaterPages=_page_decorator)
    return stream.getvalue()


__all__ = [
    "CandidateExport",
    "EXCEL_MIME",
    "EXPORT_NOTICE",
    "ExportService",
    "PDF_MIME",
    "neutralize_excel_formula",
]
